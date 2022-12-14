import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border ,Side

currency_to_rub = {"AZN": 35.68 ,"BYR": 23.91 ,"EUR": 59.90 ,"GEL": 21.74 ,"KGS": 0.76 ,"KZT": 0.13 ,"RUR": 1 ,
    "UAH": 1.64 ,"USD": 60.66 ,"UZS": 0.0055 ,}


def addWidth(work_sheet):
    dims = {}
    for row in work_sheet.rows:
        for cell in row:
            dims[cell.column_letter] = max((dims.get(cell.column_letter ,0) ,len(str(cell.value))))
    for key ,width in dims.items():
        work_sheet.column_dimensions[key].width = width + 2
    return work_sheet


def createTable(ws ,columns ,data ,precent):
    for item in columns:
        data.append(list(item))

    side = Side(style='thin')
    border = Border(left=side ,right=side ,top=side ,bottom=side)

    for row in data:
        ws.append(row)

    for index in range(len(data[0])):
        ws.cell(row=1 ,column=index + 1).font = Font(bold=True)
        ws.cell(row=1 ,column=index + 1).border = border

        for i in range(len(data) - 1):
            if precent and index == 4:
                ws.cell(row=i + 2 ,column=index + 1).value = str(
                    round(float(str(ws.cell(row=i + 2 ,column=index + 1).value)) * 100 ,2)) + '%'
            ws.cell(row=i + 2 ,column=index + 1).border = border

    return addWidth(ws)


class Report():
    def __init__(self ,year ,city ,profession):
        self.year = year
        self.city = city
        self.profession = profession

        self.wb = Workbook()

        # Первый лист
        self.ws1 = self.wb.active
        self.ws1.title = "Статистика по годам"

        # Второй лист
        self.ws2 = self.wb.create_sheet("Статистика по городам")

    def generate_excel(self):
        dynamic_salary = list(self.year[0].values())
        dynamic_salary_profession = list(self.year[2].values())
        dynamic_counts_vacancies = list(self.year[1].values())
        dynamic_counts_vacancies_profession = list(self.year[3].values())
        dynamic_years = [list(self.year[0].keys())]
        arrayforws1 = np.concatenate(
            [np.array(dynamic_years).reshape(-1 ,1)] + [np.array(dynamic_salary).reshape(-1 ,1)] + [
                np.array(dynamic_salary_profession).reshape(-1 ,1)] + [
                np.array(dynamic_counts_vacancies).reshape(-1 ,1)] + [
                np.array(dynamic_counts_vacancies_profession).reshape(-1 ,1)] ,axis=1)
        header = [['Год' ,'Средняя зарплата' ,f'Средняя зарплата - {self.profession}' ,'Количество вакансий' ,
            f'Количество вакансий - {self.profession}']]
        self.ws1 = createTable(self.ws1 ,arrayforws1 ,header ,False)

        city_salary_keys = list(self.city[0].keys())
        city_salary_values = list(self.city[0].values())
        city_precent_keys = list(self.city[1].keys())
        city_precent_values = list(self.city[1].values())
        arrayforws2 = np.concatenate(((([np.array(city_salary_keys).reshape(-1, 1)] + [np.array(city_salary_values).reshape(-1, 1)] + [np.array(['' for _ in range(len(city_salary_keys))]).reshape(-1, 1)]) + [np.array(city_precent_keys).reshape(-1, 1)]) + [np.array(city_precent_values).reshape(-1, 1)]), axis=1)

        header = [["Город" ,"Уровень зарплат" ,"" ,"Город" ,"Доля вакансий"]]
        self.ws2 = createTable(self.ws2 ,arrayforws2 ,header ,True)

        self.wb.save('report.xlsx')


class Dataset():
    def __init__(self ,file_name ,profession):
        self.file_name = file_name
        self.profession = profession
        self.data = self.getData()
        self.year_collection = []
        self.city_collection = []

    def getData(self):
        df = pd.read_csv(self.file_name)
        df.dropna(inplace=True)
        self.header = list(df)
        return df.to_numpy()

    def getDynamic(self):
        dynamic_level_salary_for_years = {}
        dynamic_counts_vacancies_for_years = {}
        dynamic_level_salary_for_profession = {}
        dynamic_counts_salary_for_profession = {}
        for item in self.data:
            item = dict(zip(self.header ,item))

            year = int(item['published_at'].split('-')[0])
            convert = currency_to_rub[item['salary_currency']]

            if year not in list(dynamic_level_salary_for_years.keys()): dynamic_level_salary_for_years[year] = []
            if year not in list(dynamic_counts_vacancies_for_years.keys()): dynamic_counts_vacancies_for_years[year] = 0
            if year not in list(dynamic_level_salary_for_profession.keys()) and profession in item['name']: dynamic_level_salary_for_profession[year] = []
            if year not in list(dynamic_counts_salary_for_profession.keys()): dynamic_counts_salary_for_profession[
                year] = 0

            dynamic_level_salary_for_years[year].append(
                (int(float(item['salary_from'])) + int(float(item['salary_to']))) // 2 * convert)
            dynamic_counts_vacancies_for_years[year] += 1

            if self.profession in item['name']:
                dynamic_level_salary_for_profession[year].append(
                    (int(float(item['salary_from'])) + int(float(item['salary_to']))) // 2 * convert)
                dynamic_counts_salary_for_profession[year] += 1

        for key ,value in dynamic_level_salary_for_years.items(): dynamic_level_salary_for_years[key] = int(
            sum(value) / len(value))
        for key ,value in dynamic_level_salary_for_profession.items(): dynamic_level_salary_for_profession[key] = int(
            sum(value) / len(value))
        if sum(
            dynamic_counts_salary_for_profession.values()) == 0: dynamic_level_salary_for_profession = dynamic_counts_salary_for_profession

        print(f'Динамика уровня зарплат по годам: {dynamic_level_salary_for_years}')
        print(f'Динамика количества вакансий по годам: {dynamic_counts_vacancies_for_years}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {dynamic_level_salary_for_profession}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {dynamic_counts_salary_for_profession}')

        self.year_collection.append(dynamic_level_salary_for_years)
        self.year_collection.append(dynamic_counts_vacancies_for_years)
        self.year_collection.append(dynamic_level_salary_for_profession)
        self.year_collection.append(dynamic_counts_salary_for_profession)

    def getCityData(self):
        city_counts_vacancies = {}
        count_vacancy_of_city = 0
        for item in self.data:
            item = dict(zip(self.header ,item))
            area_name = item['area_name']
            convert = currency_to_rub[item['salary_currency']]
            if area_name not in list(city_counts_vacancies.keys()): city_counts_vacancies[area_name] = []
            city_counts_vacancies[area_name].append(
                (int(float(item['salary_from'])) + int(float(item['salary_to']))) // 2 * convert)
            count_vacancy_of_city += 1

        salary_of_city = {}
        percent_of_city = {}

        for key ,value in city_counts_vacancies.items():
            if len(value) >= count_vacancy_of_city // 100:
                salary_of_city[key] = int(sum(value) // len(value))
                percent_of_city[key] = round(len(value) / count_vacancy_of_city ,4)
            salary_of_city = dict(sorted(salary_of_city.items() ,key=lambda item: item[1] ,reverse=True)[:10])
            percent_of_city = dict(sorted(percent_of_city.items() ,key=lambda item: item[1] ,reverse=True)[:10])

        print(f'Уровень зарплат по городам (в порядке убывания): {salary_of_city}')
        print(f'Доля вакансий по городам (в порядке убывания): {percent_of_city}')

        self.city_collection.append(salary_of_city)
        self.city_collection.append(percent_of_city)


file_name = "vacancies_by_year.csv"
profession = "Аналитик"

df = Dataset(file_name ,profession)
df.getDynamic()
df.getCityData()

report = Report(df.year_collection ,df.city_collection ,df.profession)
report.generate_excel()